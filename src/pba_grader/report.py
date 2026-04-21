"""Output Excel rekap + PDF feedback per mahasiswa."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .schema import PlagiarismFlag, SOAL_IDS, StudentResult

log = logging.getLogger(__name__)

LEVEL_COLOR = {
    "sangat_baik": "C6EFCE",
    "baik": "DCE6F1",
    "cukup": "FFEB9C",
    "kurang": "FFC7CE",
    "tidak_dijawab": "D9D9D9",
}


def write_excel_rekap(results: list[StudentResult], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: Rekap ---
    ws = wb.active
    ws.title = "Rekap"

    headers = [
        "PDF",
        "Nama",
        "NIM",
        "Versi",
        *[f"{sid}_skor" for sid in SOAL_IDS],
        "Total_AI",
        "Skor_Final",  # kolom untuk dosen override
        "Catatan_Dosen",
        "Flags",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        row = [
            r.pdf_path.name,
            r.identity.nama or "",
            r.identity.nim or "",
            r.identity.versi,
        ]
        for sid in SOAL_IDS:
            sc = r.scores.get(sid)
            row.append(round(sc.skor_0_100, 1) if sc else "")
        row.append(round(r.total, 2))
        row.append("")  # Skor_Final kosong, utk diisi manual
        row.append("")  # Catatan_Dosen
        row.append("; ".join(r.flags))
        ws.append(row)

        # Warna sel skor berdasarkan level
        current_row = ws.max_row
        for j, sid in enumerate(SOAL_IDS):
            sc = r.scores.get(sid)
            if sc:
                col = 5 + j  # PDF(1)+Nama(2)+NIM(3)+Versi(4)+soal_index
                cell = ws.cell(row=current_row, column=col)
                fill = PatternFill("solid", fgColor=LEVEL_COLOR.get(sc.level, "FFFFFF"))
                cell.fill = fill

    # Auto width
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(12, min(40, len(str(headers[col - 1])) + 4))
    ws.freeze_panes = "E2"

    # --- Sheet 2: Detail feedback per soal ---
    ws2 = wb.create_sheet("Feedback")
    ws2.append(["PDF", "Nama", "Soal", "Level", "Skor", "Feedback", "Evidensi"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in results:
        for sid in SOAL_IDS:
            sc = r.scores.get(sid)
            if not sc:
                continue
            ws2.append(
                [
                    r.pdf_path.name,
                    r.identity.nama or "",
                    sid,
                    sc.level,
                    round(sc.skor_0_100, 1),
                    sc.feedback,
                    " | ".join(sc.evidensi),
                ]
            )
    for col, width in [("A", 30), ("B", 25), ("C", 6), ("D", 14), ("E", 6), ("F", 60), ("G", 60)]:
        ws2.column_dimensions[col].width = width

    # --- Sheet 3: Meta ---
    ws3 = wb.create_sheet("Meta")
    ws3.append(["pdf", "identity_notes", "elapsed_s"])
    for r in results:
        ws3.append([r.pdf_path.name, " | ".join(r.identity.notes), round(r.elapsed_s, 1)])

    wb.save(out_path)


def write_plagiarism_report(flags: list[PlagiarismFlag], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Plagiarism"
    ws.append(["PDF A", "PDF B", "Soal", "Cosine Similarity"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for f in flags:
        ws.append([f.pdf_a.name, f.pdf_b.name, f.soal_id, round(f.cosine_similarity, 3)])
    for col, width in [("A", 30), ("B", 30), ("C", 8), ("D", 18)]:
        ws.column_dimensions[col].width = width
    wb.save(out_path)


# --- PDF feedback per mahasiswa ---

def write_student_feedback_pdf(result: StudentResult, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=A4,
        title=f"Feedback {result.identity.nama or result.pdf_path.stem}",
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h2 = styles["Heading2"]
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=9, leading=11)

    story = []
    story.append(Paragraph("Feedback UTS — AI dalam Pendidikan Bahasa Arab", h1))
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            f"<b>Nama</b>: {result.identity.nama or '-'}<br/>"
            f"<b>NIM</b>: {result.identity.nim or '-'}<br/>"
            f"<b>Versi UTS</b>: {result.identity.versi}<br/>"
            f"<b>File</b>: {result.pdf_path.name}",
            styles["BodyText"],
        )
    )
    story.append(Spacer(1, 10))

    # Tabel ringkas skor
    story.append(Paragraph("Ringkasan Skor", h2))
    table_data = [["Soal", "Level", "Skor", "Bobot", "Skor Terbobot"]]
    for sid in SOAL_IDS:
        sc = result.scores.get(sid)
        if not sc:
            continue
        table_data.append(
            [sid, sc.level, f"{sc.skor_0_100:.1f}", f"{sc.bobot}%", f"{sc.skor_terbobot:.2f}"]
        )
    table_data.append(["TOTAL", result.total_level, "-", "100%", f"{result.total:.2f}"])
    t = Table(table_data, hAlign="LEFT", colWidths=[2 * cm, 3 * cm, 2 * cm, 2 * cm, 3 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 14))

    # Feedback detail per soal
    story.append(Paragraph("Feedback Per Soal", h2))
    for sid in SOAL_IDS:
        sc = result.scores.get(sid)
        if not sc:
            continue
        story.append(
            Paragraph(
                f"<b>Soal {sid}</b> — {sc.level} ({sc.skor_0_100:.1f}/100, bobot {sc.bobot}%)",
                styles["Heading3"],
            )
        )
        fb = (sc.feedback or "-").replace("\n", "<br/>")
        story.append(Paragraph(fb, styles["BodyText"]))
        if sc.evidensi:
            story.append(Paragraph("<i>Evidensi dari jawaban:</i>", small))
            for e in sc.evidensi[:3]:
                safe = e.replace("<", "&lt;").replace(">", "&gt;")
                story.append(Paragraph(f"• {safe}", small))
        story.append(Spacer(1, 6))

    if result.flags:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Catatan Sistem", h2))
        for f in result.flags:
            story.append(Paragraph(f"• {f}", small))

    doc.build(story)
